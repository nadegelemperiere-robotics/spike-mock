# -------------------------------------------------------
# Copyright (c) [2022] Nadege LEMPERIERE
# All rights reserved
# -------------------------------------------------------
# Keywords to create data for module test
# -------------------------------------------------------
# Nadège LEMPERIERE, @1 november 2022
# Latest revision: 1 november 2022
# -------------------------------------------------------

# System includes
from math import pi

# Robotframework includes
from robot.libraries.BuiltIn import BuiltIn, _Misc
from robot.api import logger as logger
from robot.api.deco import keyword
ROBOT = False

# Openpyxl includes
from openpyxl import load_workbook

# wpimath includes
from wpimath.geometry           import Pose3d, Translation3d, Rotation3d

# Package includes
from spike.scenario.dynamics    import ScenarioDynamics
from spike.scenario.model       import ScenarioModel
from spike.scenario.parts       import ScenarioPart, ScenarioPartColorSensor, ScenarioPartMotor
from spike.scenario.ground      import ScenarioGround
from spike.scenario.parts       import ScenarioPartWheel


@keyword('Create Model')
def create_model(configuration) :

    model    = ScenarioModel()
    model.configure(configuration)
    return model

@keyword('Create Dynamics')
def create_dynamics(model, north, east, yaw) :

    dynamics = ScenarioDynamics(model)
    dynamics.configure({'north' : float(north), 'east' : float(east), 'yaw' : float(yaw) / pi * 180})
    dynamics.extrapolate(0)
    return dynamics

@keyword('Read Test Case Parameters')
def read_test_case_parameters(filename, sheetname) :

    result = {}

    wbook = load_workbook(filename, data_only = True)
    sheet = wbook[sheetname]

    # Associate header to column
    i_column = 1
    column_to_header = {}
    header_to_column = {}
    content = sheet.cell(1,i_column).value
    while content is not None :
        column_to_header[i_column]  = content
        header_to_column[content]   = i_column
        result[content] = []
        i_column += 1
        content = sheet.cell(1,i_column).value

    i_row = 2
    while sheet.cell(i_row,1).value is not None:
        for col,header in column_to_header.items() :
            value = str(sheet.cell(i_row,col).value)
            result[header].append(value)
        i_row += 1

    return result

@keyword('Create Part')
def create_part(tpe, x, y, z, roll, pitch, yaw) :

    part      = ScenarioPart()
    part.pose = Pose3d(
        Translation3d(float(x),float(y),float(z)),
        Rotation3d(float(roll)  * pi / 180, float(pitch) * pi / 180, float(yaw) * pi / 180))
    part.type = tpe

    return part

@keyword('Create ColorSensor')
def create_colorsensor(x, y, z, roll, pitch, yaw) :

    part      = ScenarioPart()
    part.id   = ScenarioPartColorSensor.s_ids[0]
    part      = ScenarioPartColorSensor(part)
    part.pose = Pose3d(
        Translation3d(float(x),float(y),float(z)),
        Rotation3d(float(roll)  * pi / 180, float(pitch) * pi / 180, float(yaw) * pi / 180))

    return part

@keyword('Create Motor')
def create_motor(x, y, z, roll, pitch, yaw) :

    part      = ScenarioPart()
    part.id   = ScenarioPartMotor.s_ids[0]
    part      = ScenarioPartMotor(part)
    part.pose = Pose3d(
        Translation3d(float(x),float(y),float(z)),
        Rotation3d(float(roll)  * pi / 180, float(pitch) * pi / 180, float(yaw) * pi / 180))

    return part

@keyword('Create Pose')
def create_pose(x, y, z, roll, pitch, yaw) :

    pose = Pose3d(
        Translation3d(float(x),float(y),float(z)),
        Rotation3d(float(roll)  * pi / 180, float(pitch) * pi / 180, float(yaw) * pi / 180))
    return pose

@keyword('Create Mat')
def create_mat(image, path) :

    mat = ScenarioGround()
    mat.configure({'image' : image, 'scale' : 0.1}, path)
    return mat
