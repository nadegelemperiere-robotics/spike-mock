# -------------------------------------------------------
# Copyright (c) [2022] Nadege LEMPERIERE
# All rights reserved
# -------------------------------------------------------
""" Scenario data management """
# -------------------------------------------------------
# Nadège LEMPERIERE, @04 november 2022
# Latest revision: 04 november 2022
# -------------------------------------------------------

# Openpyxl includes
from openpyxl import load_workbook

class ScenarioData() :
    """ Class managing simulation data """

    s_time_header       = 'time'
    s_invalid_value     = -1e16

    def __init__(self) :
        """ Contructor for each instantiation / do nothing """

        self.__data         = {}
        self.__is_loaded    = False

    def configure(self, filename, sheet) :
        """ Read scenario data from excel file

        :param filename: path to excel file in which the scenario data are located
        :type filename:  string
        :param sheet:    sheet from which data shall be retrieved
        :type sheet:     string
        """

        self.__data = {}

        # Open workbook and load sheet
        wbook = load_workbook(filename, data_only = True)
        content_sheet = wbook[sheet]

        # Associate header with data
        self.__data = self.__load_data(content_sheet)

        if not self.s_time_header in self.__data :
            raise ValueError('Time data not found')

        self.__is_loaded = True

    def extrapolate(self, header, time) :
        """
        Extrapolate a given data at a given time

        :param header: header of the data to extrapolate
        :type header:  string
        :param time:   extrapolation date in seconds
        :type time:    float

        :raise ValueError: header not found in data

        :return:       The extrapolated value at the input time
        :rtype:        depending on data type
        """

        result = self.s_invalid_value

        if not header in self.__data :
            raise ValueError('Header ' + header + 'not found in data')
        last_time = self.__data[self.s_time_header][0]
        last_data = self.__data[header][0]
        found = False
        for i_data, t_data in enumerate(self.__data[self.s_time_header]) :
            if t_data > time and not found:
                found = True
                if  isinstance(last_data, (bool, str)) or \
                    isinstance(self.__data[header][i_data], str) or \
                    last_data is None or \
                    self.__data[header][i_data] is None :
                    result = last_data
                elif isinstance(last_data, (int, float)) :
                    result = ((t_data - time) * last_data + \
                        (time - last_time) * self.__data[header][i_data]) / \
                        (t_data - last_time)

            last_time = t_data
            last_data = self.__data[header][i_data]

        return result

    def is_loaded(self) :
        """
        Loaded status return function :

        :return: True if data are loaded, False otherwise
        :rtype:  boolean
        """

        return self.__is_loaded

    def __load_data(self, sheet) :
        """
        Reads data from the excel sheet

        :param sheet: sheet to read object status from
        :type sheet:  string
        """

        result = {}

        # Associate header to column
        i_column = 1
        column_to_header = {}
        header_to_column = {}
        content = sheet.cell(1,i_column).value
        while content is not None :
            column_to_header[i_column]  = content
            header_to_column[content]   = i_column
            result[content] = []
            i_column += 1
            content = sheet.cell(1,i_column).value

        for i_row in range(2,sheet.max_row + 1) :
            for col,header in column_to_header.items() :
                value = sheet.cell(i_row,col).value
                if isinstance(value,str) and value == 'True'  : value = True
                if isinstance(value,str) and value == 'False' : value = False
                result[header].append(value)

        return result
