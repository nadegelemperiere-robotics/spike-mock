# -------------------------------------------------------
# Copyright (c) [2022] Nadege LEMPERIERE
# All rights reserved
# -------------------------------------------------------
""" Class providing abaqus for robot behaviour """
# -------------------------------------------------------
# Nadège LEMPERIERE, @08 march 2022
# Latest revision: 08 march 2022
# -------------------------------------------------------

# Openpyxl includes
from openpyxl import load_workbook

class ScenarioAbaqus() :
    """ Class storing a robot abaqus
        The abaqus is provided as an excel workbook linking an input
        to one or more values. The abaqus topics are derived from the
        workbook sheet headers
    """

    def __init__(self) :
        """ Constructor """

        # Dictionary containing per topic the correspondance between input and output values
        self.__data = {}

    def read(self, filename, sheet) :
        """ Load an abaqus from an input workbook sheet content

        :param filename: name of the workbook file to read data from
        :type filename:  string
        :param sheet:    name of the workbook sheet containing the abaqus
        :type sheet:     string

        """

         # Open workbook and load sheets
        wbook = load_workbook(filename, data_only = True)
        sheet = wbook[sheet]

        # Associate header to column
        i_row = 2
        command = sheet.cell(i_row,1).value
        while command is not None :
            i_col = 2
            while sheet.cell(i_row,i_col).value is not None :
                if sheet.cell(1,i_col).value not in self.__data :
                    self.__data[sheet.cell(1,i_col).value] = {}
                self.__data[sheet.cell(1,i_col).value][command] = sheet.cell(i_row,i_col).value
                i_col += 1
            i_row += 1
            command = sheet.cell(i_row,1).value

    def get(self, topic, command) :
        """ Get abaqus value for a given command

        :param topic:   command topic
        :type topic:    string
        :param command: command value
        :type command:  integer
        :return:        topic value for the given command.
        :rtype:         float

        """
        result = None

        if not topic in self.__data :
            raise ValueError("Unknown topic : " + topic)
        if not command in self.__data[topic] :
            raise ValueError("Unknown command : " + str(command) + ' for topic ' + topic)

        result = self.__data[topic][command]
        return result
